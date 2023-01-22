from PyQt5.QtWidgets import *
from PyQt5.QtWidgets import * 
from PyQt5.QtGui import * 
import sys
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from os import getcwd
from PyQt5 import QtWidgets
import cv2
import random

# Наследование класса PyQt5
class dlgMain(QDialog):

    def __init__(self):
        # В конструкторе перемнных через супер получаем все селфы от наследования
        super().__init__()

        self.puth = ""
        self.coord = []
        self.val = []
        self.result = []
        self.sch = "1"


        self.resize(318, 150)
        # Рисуем элементы окна
        self.button = QPushButton("Вычислить значения", self)
        self.button.setFont(QFont("Arial", 12))
        self.button.move(60, 75)
        self.button.clicked.connect(self.save_value)

        self.button = QPushButton("Загрузить изображение", self)
        self.button.setFont(QFont("Arial", 12))
        self.button.move(47, 25)
        self.button.clicked.connect(self.load_img)

    # Хендлер второй кнокпки
    def load_img(self):
        self.puth = QFileDialog.getOpenFileName(self, "open", getcwd(), "IMG (*.png *.jpg)")
        self.img = cv2.imread(self.puth[0])

        # Обнуляем параметры
        self.coord = []
        self.val = []
        self.result = []
        self.sch = "1"

        f = open('value.txt', 'w')
        f.close()

        f = open('coordinate.txt', 'w')
        f.close()
        
        # Рисуем окошко
            # self.button.setText("Вычислить")
        cv2.imshow('image', self.img)
        cv2.setMouseCallback('image', self.mouse_click)
        cv2.waitKey(0)
        cv2.destroyAllWindows()
    
    # Обработчик второй кнопки
    def save_value(self):

        if self.puth != "":
            

            with open("coordinate.txt", "r") as file:
                self.coord = file.read().split()
            
            with open("value.txt", "r") as file:
                self.val = file.read().split()

            # Я был упоротый когда писал формулу, но она работает, лучше не трогать
            difference_value = abs(int(self.val[0]) - int(self.val[1]))
            difference_coord = abs(int(self.coord[0]) - int(self.coord[1])) 

            self.point_len = difference_coord / difference_value


        for i in range(2, len(self.coord)):
            self.result.append(round(abs(int(self.coord[1]) - int(self.coord[i])) / self.point_len))
        
        for i in range(len(self.result)):

            self.result[i] = abs(int(self.val[1]) - self.result[i])
            print(self.result[i])
        
        

        self.result.insert(0, int(self.val[0]))
        self.result.append(int(self.val[1]))
        self.norm_func(self.result)


    # Записываем график в Excel
    def norm_func(self, coort):
        name = str(random.randint(1, 7634578634))
        wb = Workbook()
        wb.create_sheet(title=name, index=0)
        shhet = wb[name]

        for i in range(len(coort)):
            cell = shhet.cell(row=i + 1, column=1)

            cell.value = coort[i]
       
            print(cell.value)

        c1 = LineChart()
        c1.title = name
    
        data = Reference(shhet, min_col=1, min_row=1, max_col=1, max_row=len(coort))
        c1.add_data(data)

        shhet.add_chart(c1, 'C2')
        wb.save(name + '.xlsx')
        self.inf1()
        return name + '.xlsx'



    def mouse_click(self, event, x, y, flags, param):

        if event == cv2.EVENT_LBUTTONDOWN:
            
            font = cv2.FONT_HERSHEY_DUPLEX
            cv2.putText(self.img, self.sch, (x, y), 
                    font, 1, 
                    (0, 255, 255), 
                    2) 
            cv2.imshow('image', self.img)

            self.sch = int(self.sch) + 1
            self.sch = str(self.sch)

            with open("value.txt", "r") as file:
                spis_value = file.read().split()

            with open("coordinate.txt", "r") as file:
                spis_coord = file.read().split()	

            if len(spis_value) == 0 and len(spis_coord) == 0:
                dlgMain.takeinputs(self)

                with open("coordinate.txt", "a") as file:
                    file.write(str(y) + " ")

            if len(spis_value) == 1 and len(spis_coord) == 1:
                dlgMain.takeinputs1(self)

                with open("coordinate.txt", "a") as file:
                    file.write(str(y) + " ")

                dlgMain.inf(self)
                    
            
            if len(spis_value) >= 2 and len(spis_coord) >= 2:
                with open("coordinate.txt", "a") as file:
                    file.write(str(y) + " ")



    def takeinputs(self):
        name, done1 = QtWidgets.QInputDialog.getText(
        self, 'ТЧК', 'Введите первую точку: ')
        print(name)

        with open("value.txt", "a") as file:
            file.write(name + " ")

    def takeinputs1(self):
        name, done1 = QtWidgets.QInputDialog.getText(
        self, 'ТЧК', 'Введите последнюю точку: ')
        print(name + " ")

        with open("value.txt", "a") as file:
            file.write(name + " ")

    def inf(self):
        zzz = QMessageBox.information(self, "Информация", "Теперь нажмите на точки, значения которых необходимо вычислить")

    def inf1(self):
        z = QMessageBox.information(self, "Информация", "График успешно сохранён")

# Создаём экземляр программы и запускаем
if __name__ == "__main__":
    app = QApplication(sys.argv)

    # Создаём экземпляp созданного изменённого файла
    main = dlgMain()

    # Рисуем окно виджетов
    main.show()

    # Бесконечный цикл (луп) 

    sys.exit(app.exec_())



